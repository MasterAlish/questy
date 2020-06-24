#coding:utf-8
from copy import copy

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Side, Border, colors, Alignment

""" 
    Для методов которые принимают start_point и end_point в качестве координат можно передавать:
        Не указывая end_point: 
            одну строку - some_method(start_point=5) -единственный случай, где тип данных координаты 'int'     
            один столбец - some_method(start_point='C')     
            одну ячейку - some_method(start_point='D6')
        Указывая end_point:
            срез строк - some_method(start_point='3', end_point='8') - Тип данных 'str'
            срез столбцов - some_method(start_point='C', end_point='F') 
            произвольную область - some_method(start_point='B2', end_point='F6')
"""


class ExcelStyleMaster:
    def __init__(self,wb=Workbook()):
        self.wb = wb
        self.ws = self.wb.active

    def fill_cells(self, start_point, end_point=None, color=colors.WHITE, fill_type='solid'):
        if end_point is None:
            if type(start_point) is int:
                for cell in self.ws[start_point]:
                    cell.fill = PatternFill(fill_type=fill_type, fgColor=color)
            elif start_point.isalpha():
                for cell in self.ws[start_point]:
                    cell.fill = PatternFill(fill_type=fill_type, fgColor=color)
        else:
            rows = self.ws[start_point + ':' + end_point]
            for row in rows:
                for cell in row:
                    cell.fill = PatternFill(fill_type=fill_type, fgColor=color)

    def set_column_width(self, col, width=None, auto=False ):
        if width is not None:
            self.ws.column_dimensions[col].width = width
        elif auto:
            column = self.ws[col]
            for cell in column:
                if len(unicode(cell.value)) > self.ws.column_dimensions[col].width:
                    self.ws.column_dimensions[col].width = len(unicode(cell.value))

    def set_row_height(self, row, height):
        self.ws.row_dimensions[row].height = height

    def set_alighment(self, cell, h_align, v_align=None):
        cell = self.ws.cell(cell)
        cell.alignment = Alignment(horizontal=h_align, vertical=v_align)

    def make_border(self, start_point, end_point=None, color=colors.BLACK, border_style='thin'):# FIXME сократить if`s
        top = Border(top=Side(border_style=border_style, color=color))
        bottom = Border(bottom=Side(border_style=border_style, color=color))
        left = Border(left=Side(border_style=border_style, color=color))
        right = Border(right=Side(border_style=border_style, color=color))

        if end_point is None:
            if type(start_point) is int:
                row = self.ws[start_point]
                row[0].border = self._add_border(row[0], left)
                row[-1].border = self._add_border(row[-1], right)
                for cell in row:
                    cell.border = self._add_border(cell, top)
                    cell.border = self._add_border(cell, bottom)

            elif start_point.isalpha():
                col = self.ws[start_point]
                col[0].border = self._add_border(col[0], top)
                col[-1].border = self._add_border(col[-1], bottom)
                for cell in col:
                    cell.border = self._add_border(cell, left)
                    cell.border = self._add_border(cell, right)
            elif start_point.isalnum():
                cell = self.ws[start_point]
                cell.border = self._add_border(cell, left)
                cell.border = self._add_border(cell, right)
                cell.border = self._add_border(cell, top)
                cell.border = self._add_border(cell, bottom)

        else:
            if start_point.isdigit() or start_point.isalnum():
                rows = self.ws[start_point + ':' + end_point]
                for cell in rows[0]:
                    cell.border = self._add_border(cell, top)
                for cell in rows[-1]:
                    cell.border = self._add_border(cell, bottom)
                for row in rows:
                    row[0].border = self._add_border(row[0], left)
                    row[-1].border = self._add_border(row[0], right)
            elif start_point.isalpha():
                cols = self.ws[start_point + ':' + end_point]
                for cell in cols[0]:
                    cell.border = self._add_border(cell, left)
                for cell in cols[-1]:
                    cell.border = self._add_border(cell, right)
                for col in cols:
                    col[0].border = self._add_border(col[0], top)
                    col[-1].border = self._add_border(col[-1], bottom)

    def merge_cells(self, **coordinates):
        self.ws.merge_cells(**coordinates)

    def set_font_style(self, start_point,
                       end_point=None,
                       name=None,
                       color=None,
                       bold=None,
                       italic=None,
                       strike=None,
                       size=None,
                       underline=None):
        style_map = {'name': name, 'color': color, 'bold': bold, 'italic': italic, 'strike': strike, 'size': size, 'underline': underline}
        if end_point is None:
            if type(start_point) is int:
                row = self.ws[start_point]
                for cell in row:
                    self._add_font_style(cell, style_map)
            elif start_point.isalpha():
                col = self.ws[start_point]
                for cell in col:
                    self._add_font_style(cell, style_map)
            elif start_point.isalnum():
                cell = self.ws[start_point]
                self._add_font_style(cell, style_map)
        else:
            cells_range = self.ws[start_point + ':' + end_point]
            for row in cells_range:
                for cell in row:
                    self._add_font_style(cell, style_map)

    def _add_border(self, cell, adding_border):  # из за проблем с наложением рамок ячеек пришлось усложнить
        new_border = copy(cell.border)
        if adding_border.left != Side():
            new_border.left = adding_border.left
        if adding_border.right != Side():
            new_border.right = adding_border.right
        if adding_border.top != Side():
            new_border.top = adding_border.top
        if adding_border.bottom != Side():
            new_border.bottom = adding_border.bottom
        return new_border

    def _add_font_style(self, cell, style_map):
        global name, color, bold, italic, strike, size, underline
        font = copy(cell.font)
        if style_map['name'] is not None:
            font.name = style_map['name']
        if style_map['color'] is not None:
            font.color = style_map['color']
        if style_map['bold'] is not None:
            font.bold = style_map['bold']
        if style_map['italic'] is not None:
            font.italic = style_map['italic']
        if style_map['strike'] is not None:
            font.strike = style_map['strike']
        if style_map['size'] is not None:
            font.size = style_map['size']
        if style_map['underline'] is not None:
            font.underline = style_map['underline']
        cell.font = font

"""
FILL STYLES
    'none'
    'solid'
    'darkDown'
    'darkGray'
    'darkGrid'
    'darkHorizontal'
    'darkTrellis'
    'darkUp'
    'darkVertical'
    'gray0625'
    'gray125'
    'lightDown'
    'lightGray'
    'lightGrid'
    'lightHorizontal'
    'lightTrellis'
    'lightUp'
    'lightVertical'
    'mediumGray'

BORDER STYLES
    None
    'dashDot'
    'dashDotDot'
    'dashed'
    'dotted'
    'double'
    'hair'
    'medium'
    'mediumDashDot'
    'mediumDashDotDot'
    'mediumDashed'
    'slantDashDot'
    'thick'
    'thin'

UNDERLINE STYLES
    'single'
    'singleAccounting'
    'double'
    'doubleAccounting'
    
horizontal_alignments = (
    "general", 
    "left", 
    "center", 
    "right", 
    "fill", 
    "justify", 
    "centerContinuous",
    "distributed", )
    
vertical_aligments = (
    "top", 
    "center", 
    "bottom", 
    "justify", 
    "distributed",
)
"""