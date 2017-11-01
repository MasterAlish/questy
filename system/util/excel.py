from django.conf import settings
from openpyxl import Workbook
import datetime
import os


def test_excel():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 42
    ws.append([1, 2, 3])

    ws['A2'] = datetime.datetime.now()

    wb.save(os.path.join(settings.MEDIA_ROOT, "sample.xlsx"))